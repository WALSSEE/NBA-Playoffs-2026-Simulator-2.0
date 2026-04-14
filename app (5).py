import streamlit as st
import pandas as pd
import numpy as np
import io

st.set_page_config(page_title="NBA Playoffs 2026", page_icon="🏀", layout="wide", initial_sidebar_state="expanded")

st.markdown("""
<style>
@import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Inter:wght@300;400;500;600&display=swap');
html,body,[class*="css"]{font-family:'Inter',sans-serif;}
h1,h2,h3{font-family:'Bebas Neue',sans-serif;letter-spacing:2px;}
.stApp{background:#0a0a0f;color:#e8e8e8;}
[data-testid="stSidebar"]{background:#111118!important;border-right:1px solid #222230;}
.mc{background:linear-gradient(135deg,#1a1a2e,#16213e);border:1px solid #2a2a4a;
    border-radius:12px;padding:18px;text-align:center;margin-bottom:8px;}
.mv{font-family:'Bebas Neue',sans-serif;font-size:2.4rem;color:#f97316;line-height:1;margin:0;}
.ml{font-size:0.7rem;color:#8888aa;text-transform:uppercase;letter-spacing:1.5px;margin-top:4px;}
hr{border-color:#222230!important;}
.stButton>button{background:linear-gradient(135deg,#f97316,#ea580c)!important;color:white!important;
 border:none!important;font-family:'Bebas Neue',sans-serif!important;font-size:1rem!important;
 letter-spacing:2px!important;border-radius:8px!important;padding:8px 20px!important;}
.stTabs [data-baseweb="tab-list"]{background:#111118;border-bottom:2px solid #f97316;}
.stTabs [data-baseweb="tab"]{font-family:'Bebas Neue',sans-serif;letter-spacing:1.5px;color:#8888aa!important;}
.stTabs [aria-selected="true"]{color:#f97316!important;background:#1a1a2e!important;}
.pibox{background:linear-gradient(135deg,#1a0a2e,#0a1a2e);border:1px solid #7c3aed;
       border-radius:10px;padding:12px 16px;margin:8px 0;}
</style>
""", unsafe_allow_html=True)


# ══════════════════════════════════════════════════════════════════════════════
# MATH
# ══════════════════════════════════════════════════════════════════════════════

def win_prob(nr_a: float, nr_b: float, home_adv: float) -> float:
    """Win probability from net ratings. k=0.116 → 10pt gap ≈ 76% per game."""
    return float(np.clip(1 / (1 + np.exp(-(nr_a + home_adv - nr_b) * 0.116)), 0.001, 0.999))


def series_probs(nr_home: float, nr_away: float, home_adv: float,
                 variance: float = 0.0, best_of: int = 7) -> dict:
    """
    Exact series probs (DP). NBA 2-2-1-1-1 schedule.
    variance: adds noise to per-game probabilities (0=none, 1=high)
              models uncertainty / series randomness.
    nr_home = net rating of HIGHER seed (home court advantage team)
    """
    wn = (best_of + 1) // 2
    schedule = {1:True,2:True,3:False,4:False,5:True,6:False,7:True}

    # Base probs
    ph_base = win_prob(nr_home, nr_away,  home_adv)
    pa_base = win_prob(nr_home, nr_away, -home_adv)

    # With variance: blend toward 50% (shrinks edge)
    # variance=0 → no change; variance=1 → halfway to 50%
    ph = ph_base * (1 - variance * 0.5) + 0.5 * variance * 0.5
    pa = pa_base * (1 - variance * 0.5) + 0.5 * variance * 0.5

    states = {(0,0): 1.0}
    outcomes = {}
    for g in range(1, best_of+1):
        pw = ph if schedule[g] else pa
        ns = {}
        for (wh,wa), prob in states.items():
            for dwh,dwa,p in [(1,0,pw),(0,1,1-pw)]:
                nwh,nwa = wh+dwh, wa+dwa
                if nwh==wn or nwa==wn:
                    outcomes[(nwh,nwa)] = outcomes.get((nwh,nwa),0) + prob*p
                else:
                    ns[(nwh,nwa)] = ns.get((nwh,nwa),0) + prob*p
        states = ns

    p_home_win = sum(v for (wh,wa),v in outcomes.items() if wh==wn)
    p_away_win = sum(v for (wh,wa),v in outcomes.items() if wa==wn)
    aw = {k:v for k,v in outcomes.items() if k[0]==wn}
    bw = {k:v for k,v in outcomes.items() if k[1]==wn}

    def mgn(d,m): return sum(v for (wh,wa),v in d.items() if abs(wh-wa)>=m)

    return {
        'p_home': p_home_win, 'p_away': p_away_win, 'outcomes': outcomes,
        'p_home_m3': mgn(aw,3), 'p_home_m2': mgn(aw,2), 'p_home_m1': mgn(aw,1),
        'p_away_m3': mgn(bw,3), 'p_away_m2': mgn(bw,2), 'p_away_m1': mgn(bw,1),
    }


def playin_exact(t7, t8, t9, t10, nr, home_adv):
    """Analytical Play-In probs. Returns {team: {p7, p8, p_qualify}}"""
    p78  = win_prob(nr[t7], nr[t8],  home_adv)
    p910 = win_prob(nr[t9], nr[t10], home_adv)
    p_t8_t9  = win_prob(nr[t8], nr[t9],  home_adv)
    p_t8_t10 = win_prob(nr[t8], nr[t10], home_adv)
    p_t7_t9  = win_prob(nr[t7], nr[t9],  home_adv)
    p_t7_t10 = win_prob(nr[t7], nr[t10], home_adv)
    r = {t:{'p7':0.,'p8':0.} for t in [t7,t8,t9,t10]}
    r[t7]['p7'] += p78
    r[t8]['p8'] += p78*p910*p_t8_t9;      r[t9]['p8']  += p78*p910*(1-p_t8_t9)
    r[t8]['p8'] += p78*(1-p910)*p_t8_t10; r[t10]['p8'] += p78*(1-p910)*(1-p_t8_t10)
    r[t8]['p7'] += (1-p78)
    r[t7]['p8'] += (1-p78)*p910*p_t7_t9;      r[t9]['p8']  += (1-p78)*p910*(1-p_t7_t9)
    r[t7]['p8'] += (1-p78)*(1-p910)*p_t7_t10; r[t10]['p8'] += (1-p78)*(1-p910)*(1-p_t7_t10)
    for t in r: r[t]['p_qualify'] = r[t]['p7'] + r[t]['p8']
    return r


def sim_series(home, away, nr, home_adv, variance, rng, best_of=7):
    """Simulate one series via MC. home = higher seed."""
    wn = (best_of+1)//2
    ph = win_prob(nr[home], nr[away],  home_adv)
    pa = win_prob(nr[home], nr[away], -home_adv)
    # Apply variance (shrink toward 0.5)
    ph = ph*(1-variance*0.5) + 0.5*variance*0.5
    pa = pa*(1-variance*0.5) + 0.5*variance*0.5
    schedule = [True,True,False,False,True,False,True]
    wh = wa = 0
    for i in range(best_of):
        if rng.random() < (ph if schedule[i] else pa): wh += 1
        else: wa += 1
        if wh==wn: return home
        if wa==wn: return away
    return home


def sim_playin(t7, t8, t9, t10, nr, home_adv, rng):
    g1 = t7 if rng.random() < win_prob(nr[t7], nr[t8],  home_adv) else t8
    g1l = t8 if g1==t7 else t7
    g2 = t9 if rng.random() < win_prob(nr[t9], nr[t10], home_adv) else t10
    g3 = g1l if rng.random() < win_prob(nr[g1l], nr[g2], home_adv) else g2
    return g1, g3


def sim_full(east, west, home_adv, variance, n_sim):
    """
    Path-aware MC simulation of full bracket.
    Bracket structure from 2026 NBA Playoffs:

    WEST R1: (1)OKC vs (8)PIn, (4)LAL vs (5)HOU, (3)DEN vs (6)MIN, (2)SAS vs (7)PIn
    EAST R1: (1)DET vs (8)PIn, (4)CLE vs (5)NYK, (3)ATL vs (6)TOR, (2)BOS vs (7)PIn

    R2 re-seed: winners re-seed by original seed within bracket half
    """
    rng = np.random.default_rng(42)
    nr = {t['name']: t['nr'] for t in east+west}

    # Initialize counters
    cw_e  = {t['name']:0 for t in east}
    cw_w  = {t['name']:0 for t in west}
    nba_w = {t['name']:0 for t in east+west}
    rnd_e = {t['name']:{1:0,2:0,3:0} for t in east}
    rnd_w = {t['name']:{1:0,2:0,3:0} for t in west}
    pq_e  = {t['name']:{7:0,8:0} for t in east}
    pq_w  = {t['name']:{7:0,8:0} for t in west}

    e_seed = {t['seed']:t['name'] for t in east}
    w_seed = {t['seed']:t['name'] for t in west}

    def higher(a, b, seed_map):
        """Return (higher_seed_name, lower_seed_name)"""
        sa = seed_map.get(a, 99)
        sb = seed_map.get(b, 99)
        return (a, b) if sa < sb else (b, a)

    def play(a, b, seed_map):
        h, aw = higher(a, b, seed_map)
        return sim_series(h, aw, nr, home_adv, variance, rng)

    def sim_conf(seed_map, cw, rnd, pq, pi7, pi8, pi9, pi10):
        # Play-In
        q7, q8 = sim_playin(pi7, pi8, pi9, pi10, nr, home_adv, rng)
        pq[q7][7] += 1
        pq[q8][8] += 1

        # Build 8-team bracket with seeds
        # Original seeds stay: 1-6 direct, q7→seed7, q8→seed8
        bracket_seed = {seed_map[s]: s for s in range(1, 7)}
        bracket_seed[q7] = 7
        bracket_seed[q8] = 8

        def ps(a, b):
            h = a if bracket_seed[a] < bracket_seed[b] else b
            aw = b if h==a else a
            return sim_series(h, aw, nr, home_adv, variance, rng)

        # R1 matchups (fixed bracket halves — no re-seed):
        # Top half: 1v8, 4v5 → winner of 1v8 meets winner of 4v5
        # Bot half: 2v7, 3v6 → winner of 2v7 meets winner of 3v6
        s1, s2, s3, s4 = seed_map[1], seed_map[2], seed_map[3], seed_map[4]
        s5, s6 = seed_map[5], seed_map[6]

        w1v8 = ps(s1, q8)
        w4v5 = ps(s4, s5)
        w2v7 = ps(s2, q7)
        w3v6 = ps(s3, s6)

        for w in [w1v8, w4v5, w2v7, w3v6]:
            rnd[w][1] += 1

        # R2: top half vs top half, bot half vs bot half
        wA = ps(w1v8, w4v5)
        wB = ps(w2v7, w3v6)
        rnd[wA][2] += 1
        rnd[wB][2] += 1

        # Conf Final
        cf = ps(wA, wB)
        rnd[cf][3] += 1
        cw[cf] += 1
        return cf

    for _ in range(n_sim):
        e_cf = sim_conf(e_seed, cw_e, rnd_e, pq_e,
                        e_seed[7], e_seed[8], e_seed[9], e_seed[10])
        w_cf = sim_conf(w_seed, cw_w, rnd_w, pq_w,
                        w_seed[7], w_seed[8], w_seed[9], w_seed[10])
        # Finals: higher NR gets home court
        if nr[e_cf] >= nr[w_cf]:
            champ = sim_series(e_cf, w_cf, nr, home_adv, variance, rng)
        else:
            champ = sim_series(w_cf, e_cf, nr, home_adv, variance, rng)
        nba_w[champ] += 1

    return cw_e, cw_w, nba_w, rnd_e, rnd_w, pq_e, pq_w


def pct(v, n=1): return f"{v*100:.{n}f}%"


# ══════════════════════════════════════════════════════════════════════════════
# 2026 BRACKET DEFAULTS
# ══════════════════════════════════════════════════════════════════════════════
# WEST: 1.OKC 2.Spurs 3.Nuggets 4.Lakers 5.Rockets 6.Timberwolves
#       Play-In: 7.Suns 8.Trail Blazers 9.Clippers 10.Warriors
# EAST: 1.Pistons 2.Celtics 3.Hawks 4.Cavaliers 5.Knicks 6.Raptors
#       Play-In: 7.76ers 8.Magic 9.Hornets 10.Heat

DEFAULTS = {
    "west": [
        ("Oklahoma City Thunder",  1, 0.0),
        ("San Antonio Spurs",      2, 0.0),
        ("Denver Nuggets",         3, 0.0),
        ("Los Angeles Lakers",     4, 0.0),
        ("Houston Rockets",        5, 0.0),
        ("Minnesota Timberwolves", 6, 0.0),
        ("Phoenix Suns",           7, 0.0),
        ("Portland Trail Blazers", 8, 0.0),
        ("LA Clippers",            9, 0.0),
        ("Golden State Warriors", 10, 0.0),
    ],
    "east": [
        ("Detroit Pistons",        1, 0.0),
        ("Boston Celtics",         2, 0.0),
        ("Atlanta Hawks",          3, 0.0),
        ("Cleveland Cavaliers",    4, 0.0),
        ("New York Knicks",        5, 0.0),
        ("Toronto Raptors",        6, 0.0),
        ("Philadelphia 76ers",     7, 0.0),
        ("Orlando Magic",          8, 0.0),
        ("Charlotte Hornets",      9, 0.0),
        ("Miami Heat",            10, 0.0),
    ],
}


# ══════════════════════════════════════════════════════════════════════════════
# SIDEBAR
# ══════════════════════════════════════════════════════════════════════════════
with st.sidebar:
    st.markdown("## ⚙️ Malliasetukset")

    home_adv = st.slider(
        "🏠 Kotietu (NR-pistettä)", 0.0, 5.0, 3.0, 0.25,
        help="NBA historiallinen kotietu ≈ 2.5–3.5 pt. Vaikuttaa jokaiseen kotipeliin."
    )
    variance = st.slider(
        "🎲 Sarjan varianssi", 0.0, 1.0, 0.0, 0.05,
        help="0 = puhdas NR-malli. 0.5 = paljon satunnaisuutta (playoff-epävarmuus)."
    )
    n_sim = st.select_slider(
        "🔁 Simulaatiot",
        options=[10_000, 50_000, 100_000, 200_000],
        value=100_000,
        format_func=lambda x: f"{x:,}"
    )

    st.divider()
    st.markdown("### 📊 Net Rating")
    st.markdown("""
Syötä joukkueen **Net Rating** (OffRtg − DefRtg).

Lähteet:
- [NBA.com Advanced Stats](https://www.nba.com/stats/teams/advanced)
- [Basketball Reference](https://www.basketball-reference.com/leagues/NBA_2026_ratings.html)
- [Cleaning the Glass](https://cleaningtheglass.com)

| NR-ero | Pelin win% |
|--------|-----------|
| 0 pt | 58.6% (kotietu) |
| 3 pt | 64.8% |
| 5 pt | 69.5% |
| 10 pt | 79.8% |
""")

    st.divider()
    st.markdown("### 🎲 Varianssi selitys")
    st.markdown("""
**0.0** – Net rating ratkaisee täysin  
**0.3** – Kohtalainen playoff-epävarmuus  
**0.5** – Korkea satunnaisuus  
**1.0** – Lähes kolikko

Varianssi kutistaa jokaisen pelin voittotodennäköisyyttä kohti 50%:ia, mallintaen playoff-sarjojen arvaamattomuutta.
""")


# ══════════════════════════════════════════════════════════════════════════════
# HEADER
# ══════════════════════════════════════════════════════════════════════════════
st.markdown("""
<div style="background:linear-gradient(135deg,#0f0f1a,#1a0a2e,#0f1a0a);
 border:1px solid #f97316;border-radius:12px;padding:20px 28px;margin-bottom:18px;text-align:center;">
 <h1 style="font-size:2.6rem;margin:0;color:#f97316;">🏀 NBA PLAYOFFS 2026</h1>
 <p style="color:#8888aa;margin:6px 0 0;font-size:0.82rem;letter-spacing:2px;">
  PATH-AWARE MONTE CARLO · NET RATING · PLAY-IN · VARIANSSI · KOTIETU
 </p>
</div>
""", unsafe_allow_html=True)

tab_bracket, tab_series, tab_playin, tab_path, tab_excel = st.tabs([
    "🏆 BRACKET & MESTARUUS", "🎯 SARJA-ANALYYSI", "🔮 PLAY-IN", "📈 REITTIANALYYSI", "📊 EXCEL"
])


# ══════════════════════════════════════════════════════════════════════════════
# TAB 1 – BRACKET
# ══════════════════════════════════════════════════════════════════════════════

# ── localStorage bridge: lue tallennetut ratingit selaimen muistista ──────────
# Injektoidaan JS joka lukee localStorage:sta ja lähettää datan Streamlitille
# query-parametrien kautta sivun latauksen yhteydessä.
_ls_read = """
<script>
(function() {
    const saved = localStorage.getItem('nba_ratings_2026');
    if (!saved) return;
    // Jos URL:ssa ei ole parametreja, lisätään ne jotta Streamlit saa datan
    const url = new URL(window.location.href);
    if (!url.searchParams.has('_nr_loaded')) {
        url.searchParams.set('nba_ratings', saved);
        url.searchParams.set('_nr_loaded', '1');
        window.location.replace(url.toString());
    }
})();
</script>
"""

_ls_save = """
<script>
function saveRatings(data) {
    localStorage.setItem('nba_ratings_2026', JSON.stringify(data));
}
// Expose to Streamlit custom component calls
window._saveRatings = saveRatings;
</script>
"""

st.components.v1.html(_ls_save, height=0)

# ── Lataa query params → session_state jos löytyy ────────────────────────────
import json as _json
_qp = st.query_params
if "nba_ratings" in _qp and "_ratings_loaded" not in st.session_state:
    try:
        _saved = _json.loads(_qp["nba_ratings"])
        for _k, _v in _saved.items():
            st.session_state[_k] = _v
        st.session_state["_ratings_loaded"] = True
        # Puhdista URL
        _clean_params = {k:v for k,v in _qp.items()
                         if k not in ("nba_ratings","_nr_loaded")}
        st.query_params.clear()
        for k,v in _clean_params.items():
            st.query_params[k] = v
    except Exception:
        pass

# ── Alusta session_state oletuksilla jos avaimia ei ole ──────────────────────
for _ck in ("west", "east"):
    for _i, (_dn, _ds, _dnr) in enumerate(DEFAULTS[_ck]):
        if f"{_ck}{_i}n" not in st.session_state:
            st.session_state[f"{_ck}{_i}n"] = _dn
        if f"{_ck}{_i}s" not in st.session_state:
            st.session_state[f"{_ck}{_i}s"] = _ds
        if f"{_ck}{_i}r" not in st.session_state:
            st.session_state[f"{_ck}{_i}r"] = _dnr

with tab_bracket:

    # ── Team input form ──
    def team_inputs(conf_key, conf_label):
        defaults = DEFAULTS[conf_key]
        teams = []
        st.markdown(f"**{conf_label}**")
        hc = st.columns([3,1,2])
        hc[0].markdown("Joukkue"); hc[1].markdown("Sija"); hc[2].markdown("Net Rating")
        for i,(dname,dseed,dnr) in enumerate(defaults):
            if dseed == 7:
                st.markdown('<div style="border-top:1px dashed #7c3aed;margin:3px 0;color:#a78bfa;'
                            'font-size:0.7rem;letter-spacing:1px;padding-top:3px;">▼ PLAY-IN (7–10)</div>',
                            unsafe_allow_html=True)
            c1,c2,c3 = st.columns([3,1,2])
            nm = c1.text_input("n",  key=f"{conf_key}{i}n", label_visibility="collapsed")
            sd = c2.number_input("s", key=f"{conf_key}{i}s", label_visibility="collapsed",
                                 min_value=1, max_value=10)
            nr = c3.number_input("r", key=f"{conf_key}{i}r", label_visibility="collapsed",
                                 step=0.1, format="%.1f")
            teams.append({"name":nm,"seed":sd,"nr":nr})
        return teams

    cw, ce = st.columns(2)
    with cw:
        west = team_inputs("west", "🔵 Läntinen konferenssi")
    with ce:
        east = team_inputs("east", "🟠 Itäinen konferenssi")

    # ── Tallenna + Simuloi napit ──────────────────────────────────────────────
    btn_cols = st.columns([1, 2])
    with btn_cols[0]:
        if st.button("💾 TALLENNA RATINGIT", use_container_width=True):
            # Kerää kaikki NR-arvot talteen
            _data_to_save = {
                k: st.session_state[k]
                for k in st.session_state
                if any(k.startswith(p) for p in
                       ["west0","west1","west2","west3","west4","west5","west6","west7","west8","west9",
                        "east0","east1","east2","east3","east4","east5","east6","east7","east8","east9"])
            }
            _js_save = f"""
            <script>
            localStorage.setItem('nba_ratings_2026', JSON.stringify({_json.dumps(_data_to_save)}));
            </script>
            """
            st.components.v1.html(_js_save, height=0)
            st.success("✅ Ratingit tallennettu selaimen muistiin! Pysyvät tallessa sivun sulkemisen jälkeen.")
    with btn_cols[1]:
        run = st.button("🏆 SIMULOI KOKO BRACKET", use_container_width=True)

    if run:
        with st.spinner(f"Simuloidaan {n_sim:,} täyttä playoff-kautta…"):
            res = sim_full(east, west, home_adv, variance, n_sim)
        st.session_state["res"] = res
        st.session_state["east"] = east
        st.session_state["west"] = west
        st.session_state["n_sim"] = n_sim

    if "res" not in st.session_state:
        st.info("Syötä net ratingit ja paina **SIMULOI KOKO BRACKET**.")
    else:
        cw_e, cw_w, nba_w, rnd_e, rnd_w, pq_e, pq_w = st.session_state["res"]
        east_s  = st.session_state["east"]
        west_s  = st.session_state["west"]
        n_s     = st.session_state["n_sim"]
        e_seed  = {t['seed']:t['name'] for t in east_s}
        w_seed  = {t['seed']:t['name'] for t in west_s}

        # Play-In summary
        st.divider()
        st.markdown("#### 🔮 Play-In – Pääsy playoffseihin")
        pc_w, pc_e = st.columns(2)
        for col, teams_s, pq_map, label in [(pc_w,west_s,pq_w,"Länsi"),(pc_e,east_s,pq_e,"Itä")]:
            with col:
                st.markdown(f"**{label}**")
                rows = []
                for t in sorted(teams_s, key=lambda x:x['seed']):
                    if t['seed'] < 7: continue
                    q = pq_map[t['name']]
                    rows.append({"Joukkue":t['name'],"Sija":f"#{t['seed']}",
                                 "NR":f"{t['nr']:+.1f}",
                                 "Sija #7":pct(q[7]/n_s),"Sija #8":pct(q[8]/n_s),
                                 "Pääsee":pct((q[7]+q[8])/n_s)})
                st.dataframe(pd.DataFrame(rows),use_container_width=True,hide_index=True)

        # Main results table
        st.divider()
        st.markdown("#### 🏆 Bracket-kierrostilastot + mestaruus")
        all_rows = []
        for t in west_s:
            rw = rnd_w[t['name']]
            all_rows.append({
                "Joukkue": t['name']+(" 🔮" if t['seed']>=7 else ""),
                "Konf.":"Länsi","Sija":t['seed'],"NR":f"{t['nr']:+.1f}",
                "R1":pct(rw[1]/n_s),"R2":pct(rw[2]/n_s),"CF":pct(rw[3]/n_s),
                "Konf.Mestari":pct(cw_w[t['name']]/n_s),
                "NBA Mestari":pct(nba_w[t['name']]/n_s),
                "_nba":nba_w[t['name']]
            })
        for t in east_s:
            rw = rnd_e[t['name']]
            all_rows.append({
                "Joukkue": t['name']+(" 🔮" if t['seed']>=7 else ""),
                "Konf.":"Itä","Sija":t['seed'],"NR":f"{t['nr']:+.1f}",
                "R1":pct(rw[1]/n_s),"R2":pct(rw[2]/n_s),"CF":pct(rw[3]/n_s),
                "Konf.Mestari":pct(cw_e[t['name']]/n_s),
                "NBA Mestari":pct(nba_w[t['name']]/n_s),
                "_nba":nba_w[t['name']]
            })
        df_all = pd.DataFrame(all_rows).sort_values("_nba",ascending=False)
        st.dataframe(df_all[["Joukkue","Konf.","Sija","NR","R1","R2","CF","Konf.Mestari","NBA Mestari"]],
                     use_container_width=True, hide_index=True)

        top10 = df_all.nlargest(12,"_nba")
        st.bar_chart(top10.set_index("Joukkue")["_nba"].apply(lambda x: round(x/n_s*100,1)),
                     color="#f97316", height=280)

        # Per-conference
        st.divider()
        rc_w, rc_e = st.columns(2)
        for col, teams_s, cw_map, rnd_map, label, color in [
            (rc_w, west_s, cw_w, rnd_w, "Länsi", "#3b82f6"),
            (rc_e, east_s, cw_e, rnd_e, "Itä",   "#f97316"),
        ]:
            with col:
                st.markdown(f"**{label}inen konferenssi**")
                rows = []
                for t in sorted(teams_s, key=lambda x:x['seed']):
                    rw = rnd_map[t['name']]
                    rows.append({
                        "Joukkue":t['name']+(" 🔮" if t['seed']>=7 else ""),
                        "NR":f"{t['nr']:+.1f}",
                        "R1":pct(rw[1]/n_s),"R2":pct(rw[2]/n_s),"CF":pct(rw[3]/n_s),
                        "Mestari":pct(cw_map[t['name']]/n_s),
                        "_m":cw_map[t['name']]
                    })
                df_c = pd.DataFrame(rows)
                st.dataframe(df_c[["Joukkue","NR","R1","R2","CF","Mestari"]],
                             use_container_width=True, hide_index=True)
                st.bar_chart(df_c.set_index("Joukkue")["_m"].apply(lambda x:round(x/n_s*100,1)),
                             color=color, height=180)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 2 – SINGLE SERIES
# ══════════════════════════════════════════════════════════════════════════════
with tab_series:
    st.markdown("### Kahden joukkueen sarja-analyysi")
    c1,c2 = st.columns(2)
    with c1:
        st.markdown('<div style="color:#f97316;font-family:Bebas Neue,sans-serif;font-size:1.1rem;letter-spacing:1px;">🟠 KORKEAMPI SIJOITUS – KOTIETU</div>',unsafe_allow_html=True)
        sn_a = st.text_input("Joukkue A", "Oklahoma City Thunder", key="sa_n")
        nr_a = st.number_input("Net Rating A", value=12.8, step=0.1, format="%.1f", key="sa_nr")
    with c2:
        st.markdown('<div style="color:#3b82f6;font-family:Bebas Neue,sans-serif;font-size:1.1rem;letter-spacing:1px;">🔵 MATALAMPI SIJOITUS – VIERASJOUKKUE</div>',unsafe_allow_html=True)
        sn_b = st.text_input("Joukkue B", "Los Angeles Lakers", key="sb_n")
        nr_b = st.number_input("Net Rating B", value=3.8, step=0.1, format="%.1f", key="sb_nr")

    bo = st.radio("Sarjamuoto", [5,7], index=1, horizontal=True, format_func=lambda x:f"Best-of-{x}")

    if st.button("🔢 LASKE", use_container_width=True, key="calc_series"):
        res_s = series_probs(nr_a, nr_b, home_adv, variance, bo)
        wn = (bo+1)//2

        st.divider()
        ph = win_prob(nr_a, nr_b,  home_adv)
        pa = win_prob(nr_a, nr_b, -home_adv)
        st.markdown("#### Yhden pelin voittotodennäköisyys")
        mc1,mc2,mc3,mc4 = st.columns(4)
        mc1.markdown(f'<div class="mc"><p class="mv">{ph*100:.1f}%</p><p class="ml">{sn_a} kotona</p></div>',unsafe_allow_html=True)
        mc2.markdown(f'<div class="mc"><p class="mv" style="color:#3b82f6">{(1-ph)*100:.1f}%</p><p class="ml">{sn_b} vieraana</p></div>',unsafe_allow_html=True)
        mc3.markdown(f'<div class="mc"><p class="mv" style="color:#3b82f6">{(1-pa)*100:.1f}%</p><p class="ml">{sn_b} kotona</p></div>',unsafe_allow_html=True)
        mc4.markdown(f'<div class="mc"><p class="mv">{pa*100:.1f}%</p><p class="ml">{sn_a} vieraana</p></div>',unsafe_allow_html=True)

        st.divider()
        st.markdown("#### Sarjan voittotodennäköisyys")
        bha, bba = res_s['p_home']*100, res_s['p_away']*100
        ca,cb = st.columns(2)
        ca.markdown(f'<div class="mc"><p class="mv">{bha:.1f}%</p><p class="ml">{sn_a} voittaa sarjan</p></div>',unsafe_allow_html=True)
        cb.markdown(f'<div class="mc"><p class="mv" style="color:#3b82f6">{bba:.1f}%</p><p class="ml">{sn_b} voittaa sarjan</p></div>',unsafe_allow_html=True)
        st.markdown(f"""<div style="display:flex;gap:3px;margin:8px 0;">
            <div style="width:{bha:.1f}%;background:linear-gradient(90deg,#f97316,#fb923c);
             border-radius:6px 0 0 6px;padding:5px 10px;color:white;font-weight:700;
             font-size:0.8rem;white-space:nowrap;overflow:hidden;">{sn_a} {bha:.1f}%</div>
            <div style="width:{bba:.1f}%;background:linear-gradient(90deg,#3b82f6,#60a5fa);
             border-radius:0 6px 6px 0;padding:5px 10px;color:white;font-weight:700;
             font-size:0.8rem;text-align:right;white-space:nowrap;overflow:hidden;">{bba:.1f}% {sn_b}</div>
        </div>""",unsafe_allow_html=True)

        st.divider()
        st.markdown("#### Marginaalilinjat")
        mr = [{"Linja":lbl,
               f"{sn_a} kattaa":pct(res_s[f'p_home_m{m}']),
               f"{sn_b} kattaa":pct(res_s[f'p_away_m{m}'])}
              for m,lbl in [(3,"+3.5"),(2,"+2.5"),(1,"+1.5")]]
        st.dataframe(pd.DataFrame(mr).set_index("Linja"),use_container_width=True)

        st.divider()
        st.markdown("#### Sarjan lopputulokset")
        oc = []
        for (wh,wa),prob in sorted(res_s['outcomes'].items(),key=lambda x:-x[1]):
            winner = sn_a if wh==wn else sn_b
            loser  = sn_b if wh==wn else sn_a
            score  = f"{wh}–{wa}" if wh==wn else f"{wa}–{wh}"
            oc.append({"Tulos":f"{winner} {score} {loser}","Voittaja":winner,
                       "Tn":pct(prob),"_p":round(prob*100,2)})
        df_oc = pd.DataFrame(oc)
        st.dataframe(df_oc[["Tulos","Voittaja","Tn"]],use_container_width=True,hide_index=True)
        st.bar_chart(df_oc.set_index("Tulos")["_p"],color="#f97316",height=230)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 3 – PLAY-IN
# ══════════════════════════════════════════════════════════════════════════════
with tab_playin:
    st.markdown("### Play-In turnauksen analyysi")
    st.markdown("""<div class="pibox">
<b style="color:#a78bfa;font-family:Bebas Neue,sans-serif;letter-spacing:2px;">🔮 PLAY-IN RAKENNE</b><br><br>
<b>Peli 1:</b> #7 vs #8 (kotietu #7) → <b>Voittaja = playoff sija #7</b><br>
<b>Peli 2:</b> #9 vs #10 (kotietu #9) → <b>Häviäjä putoaa kaudelta</b><br>
<b>Peli 3:</b> Häviäjä(P1) vs Voittaja(P2), kotietu häviäjälle(P1) → <b>Voittaja = playoff sija #8</b>
</div>""",unsafe_allow_html=True)

    pi_conf = st.radio("Konferenssi",["Länsi","Itä"],horizontal=True,key="pi_conf")
    pi_defs = [(n,nr) for n,s,nr in DEFAULTS["west" if pi_conf=="Länsi" else "east"] if s>=7]

    pi_cols = st.columns(4)
    pi_names, pi_nrs = [],[]
    for i,(col,(dn,dnr)) in enumerate(zip(pi_cols,pi_defs)):
        with col:
            st.markdown(f'<div style="color:#a78bfa;font-family:Bebas Neue,sans-serif;">SIJA #{i+7}</div>',unsafe_allow_html=True)
            n  = col.text_input("",value=dn, key=f"pi_n{i}{pi_conf}",label_visibility="collapsed")
            nr = col.number_input("",value=dnr,step=0.1,format="%.1f",key=f"pi_nr{i}{pi_conf}",label_visibility="collapsed")
            pi_names.append(n); pi_nrs.append(nr)

    if st.button("🔮 LASKE PLAY-IN", use_container_width=True):
        pm = {pi_names[i]:pi_nrs[i] for i in range(4)}
        t7,t8,t9,t10 = pi_names
        pr = playin_exact(t7,t8,t9,t10,pm,home_adv)

        g1 = win_prob(pm[t7],pm[t8],home_adv)
        g2 = win_prob(pm[t9],pm[t10],home_adv)

        st.divider()
        st.markdown("#### Yksittäisten pelien todennäköisyydet")
        gc1,gc2,gc3 = st.columns(3)
        gc1.markdown(f'<div class="mc"><p style="color:#a78bfa;font-family:Bebas Neue,sans-serif;margin:0 0 3px;">PELI 1</p><p style="margin:2px 0;font-size:0.77rem;">{t7} vs {t8}</p><p class="mv">{g1*100:.1f}%</p><p class="ml">{t7} voittaa</p></div>',unsafe_allow_html=True)
        gc2.markdown(f'<div class="mc"><p style="color:#a78bfa;font-family:Bebas Neue,sans-serif;margin:0 0 3px;">PELI 2</p><p style="margin:2px 0;font-size:0.77rem;">{t9} vs {t10}</p><p class="mv">{g2*100:.1f}%</p><p class="ml">{t9} voittaa</p></div>',unsafe_allow_html=True)
        gc3.markdown(f'<div class="mc"><p style="color:#a78bfa;font-family:Bebas Neue,sans-serif;margin:0 0 3px;">PELI 3</p><p style="margin:2px 0;font-size:0.77rem;">Riippuu P1 & P2</p><p class="mv" style="font-size:1.6rem;">4 sk.</p><p class="ml">Katso alta</p></div>',unsafe_allow_html=True)

        st.divider()
        st.markdown("#### Pääsy playoffseihin")
        pi_rows = [{"Joukkue":n,"Alkup.sija":f"#{i+7}","NR":f"{pi_nrs[i]:+.1f}",
                    "Sija #7":pct(pr[n]['p7']),"Sija #8":pct(pr[n]['p8']),
                    "Pääsee":pct(pr[n]['p_qualify']),"_q":pr[n]['p_qualify']}
                   for i,n in enumerate(pi_names)]
        df_pi = pd.DataFrame(pi_rows).sort_values("_q",ascending=False)
        st.dataframe(df_pi[["Joukkue","Alkup.sija","NR","Sija #7","Sija #8","Pääsee"]],
                     use_container_width=True,hide_index=True)

        st.markdown("#### Visualisointi")
        for _,row in df_pi.iterrows():
            nm = row["Joukkue"]
            p7  = pr[nm]['p7']*100; p8 = pr[nm]['p8']*100
            po  = (1-pr[nm]['p_qualify'])*100
            st.markdown(f"**{nm}** ({row['Alkup.sija']})")
            st.markdown(f"""<div style="display:flex;gap:2px;margin:2px 0 8px;">
                <div style="width:{p7:.1f}%;background:linear-gradient(90deg,#f97316,#fb923c);
                 border-radius:5px 0 0 5px;padding:3px 7px;color:white;font-size:0.72rem;
                 font-weight:700;white-space:nowrap;overflow:hidden;min-width:0;">Sija #7: {p7:.1f}%</div>
                <div style="width:{p8:.1f}%;background:linear-gradient(90deg,#7c3aed,#a78bfa);
                 padding:3px 7px;color:white;font-size:0.72rem;font-weight:700;
                 white-space:nowrap;overflow:hidden;min-width:0;">Sija #8: {p8:.1f}%</div>
                <div style="width:{po:.1f}%;background:#1a1a2e;border-radius:0 5px 5px 0;
                 padding:3px 7px;color:#8888aa;font-size:0.72rem;
                 white-space:nowrap;overflow:hidden;min-width:0;">Putoaa: {po:.1f}%</div>
            </div>""",unsafe_allow_html=True)

        st.divider()
        st.markdown("#### Peli 3 – kaikki skenaariot")
        p78v  = win_prob(pm[t7],pm[t8],home_adv)
        p910v = win_prob(pm[t9],pm[t10],home_adv)
        sc_rows=[]
        for (g1w,g1l),(g2w,_),ps_ in [
            ((t7,t8),(t9,t10), p78v*p910v),
            ((t7,t8),(t10,t9), p78v*(1-p910v)),
            ((t8,t7),(t9,t10),(1-p78v)*p910v),
            ((t8,t7),(t10,t9),(1-p78v)*(1-p910v)),
        ]:
            pg3 = win_prob(pm[g1l],pm[g2w],home_adv)
            sc_rows.append({"Skenaario":f"P1: {g1w} voittaa, P2: {g2w} voittaa → {g1l} vs {g2w}",
                            "Tn":pct(ps_),f"{g1l} voittaa G3":pct(pg3),f"{g2w} voittaa G3":pct(1-pg3)})
        st.dataframe(pd.DataFrame(sc_rows),use_container_width=True,hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 4 – PATH ANALYSIS
# ══════════════════════════════════════════════════════════════════════════════
with tab_path:
    st.markdown("### Reittianalyysi")
    st.caption("Simuloi ensin Bracket & Mestaruus -välilehdellä.")

    if "res" not in st.session_state:
        st.info("Aja simulaatio ensin.")
    else:
        cw_e, cw_w, nba_w, rnd_e, rnd_w, pq_e, pq_w = st.session_state["res"]
        east_s = st.session_state["east"]
        west_s = st.session_state["west"]
        n_s    = st.session_state["n_sim"]
        all_t  = east_s + west_s

        sel = st.selectbox("Valitse joukkue",
                           sorted([t['name'] for t in all_t], key=lambda x:-nba_w[x]))

        t_obj  = next(t for t in all_t if t['name']==sel)
        is_e   = t_obj in east_s
        rnd_m  = rnd_e if is_e else rnd_w
        cw_m   = cw_e  if is_e else cw_w
        pq_m   = pq_e  if is_e else pq_w
        rw     = rnd_m[sel]
        playin = t_obj['seed'] >= 7

        st.divider()
        st.markdown(f"#### {sel} – todennäköisyydet vaiheittain")

        stages = []
        if playin:
            q = pq_m[sel]
            stages += [
                ("🔮 Play-In: pääsee playoffseihin", (q[7]+q[8])/n_s, "#7c3aed"),
                ("🔮 Play-In: sija #7",               q[7]/n_s,        "#a78bfa"),
                ("🔮 Play-In: sija #8",               q[8]/n_s,        "#6d28d9"),
            ]
        stages += [
            ("✅ Voittaa 1. kierroksen",              rw[1]/n_s, "#f97316"),
            ("✅ Voittaa 2. kierroksen (semifinaali)", rw[2]/n_s, "#fb923c"),
            ("✅ Voittaa konferenssifinalin",          rw[3]/n_s, "#fbbf24"),
            ("🏆 Konferenssimestari",                  cw_m[sel]/n_s,   "#22c55e"),
            ("🏆 NBA-mestari",                         nba_w[sel]/n_s,  "#16a34a"),
        ]

        for label, prob, color in stages:
            p = prob*100
            st.markdown(f"**{label}**")
            st.markdown(f"""<div style="display:flex;align-items:center;gap:10px;margin:1px 0 9px;">
                <div style="flex:1;background:#1a1a2e;border-radius:6px;overflow:hidden;height:24px;">
                    <div style="width:{min(p,100):.1f}%;background:{color};height:100%;
                     display:flex;align-items:center;padding-left:10px;color:white;
                     font-weight:700;font-size:0.78rem;white-space:nowrap;">{p:.1f}%</div>
                </div>
            </div>""",unsafe_allow_html=True)

        # Conditional probabilities
        st.divider()
        st.markdown("#### Ehdolliset todennäköisyydet")
        st.caption("Jos joukkue on jo edennyt tähän vaiheeseen, mikä on seuraavan voittamisen tn%?")
        cond_rows = []
        prev = 1.0
        for label, prob, _ in stages:
            if "Play-In: sija" in label: continue
            cond = (prob / prev) if prev > 0.001 else 0
            cond_rows.append({"Vaihe":label, "Absoluuttinen":f"{prob*100:.1f}%",
                               "Ehdollinen":f"{cond*100:.1f}%"})
            prev = prob
        st.dataframe(pd.DataFrame(cond_rows), use_container_width=True, hide_index=True)

        # NR comparison vs all possible opponents
        st.divider()
        st.markdown("#### Sarjavoittotodennäköisyys kaikkia vastustajia vastaan")
        nr_t    = t_obj['nr']
        others  = [t for t in all_t if t['name']!=sel]
        vs_rows = []
        for opp in sorted(others, key=lambda x:-x['nr']):
            same = (opp in east_s) == is_e
            if nr_t >= opp['nr']:
                r = series_probs(nr_t, opp['nr'], home_adv, variance, 7)
                p_ser = r['p_home']
            else:
                r = series_probs(opp['nr'], nr_t, home_adv, variance, 7)
                p_ser = r['p_away']
            vs_rows.append({"Vastustaja":opp['name'],
                            "Konf.":"Sama" if same else "Finals",
                            "Vast. NR":f"{opp['nr']:+.1f}",
                            "NR ero":f"{nr_t-opp['nr']:+.1f}",
                            "Sarjavoitto":pct(p_ser)})
        st.dataframe(pd.DataFrame(vs_rows), use_container_width=True, hide_index=True)


# ══════════════════════════════════════════════════════════════════════════════
# TAB 5 – EXCEL TEMPLATE
# ══════════════════════════════════════════════════════════════════════════════
with tab_excel:
    st.markdown("### 📊 Excel-pohja – Net Rating")

    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = openpyxl.Workbook()
    ws = wb.active; ws.title = "Net Ratings 2026"

    hdrs = ["Joukkue","Konferenssi","Sija","Net Rating","Huomio"]
    hfill = PatternFill("solid",fgColor="1a1a2e")
    hfont = Font(bold=True,color="F97316",name="Calibri",size=12)
    thin  = Side(border_style="thin",color="2a2a4a")
    brd   = Border(left=thin,right=thin,top=thin,bottom=thin)
    for ci,h in enumerate(hdrs,1):
        c = ws.cell(1,ci,h); c.font=hfont; c.fill=hfill
        c.alignment=Alignment(horizontal="center"); c.border=brd

    all20 = [(n,"Länsi",s,nr,"") for n,s,nr in DEFAULTS["west"]] + \
            [(n,"Itä",  s,nr,"") for n,s,nr in DEFAULTS["east"]]

    dfont = Font(name="Calibri",size=11,color="E8E8E8")
    for ri,row in enumerate(all20,2):
        fill = PatternFill("solid",fgColor="1a0a2e") if row[2]>=7 else \
               PatternFill("solid",fgColor="0a0a0f" if ri%2==0 else "111118")
        for ci,val in enumerate(row,1):
            c=ws.cell(ri,ci,val); c.font=dfont; c.fill=fill
            c.alignment=Alignment(horizontal="center" if ci>1 else "left"); c.border=brd

    ws.column_dimensions['A'].width=28; ws.column_dimensions['B'].width=14
    ws.column_dimensions['C'].width=8;  ws.column_dimensions['D'].width=14
    ws.column_dimensions['E'].width=20

    buf=io.BytesIO(); wb.save(buf); buf.seek(0)
    st.download_button("⬇️ LATAA EXCEL-POHJA (2026 bracket)",
        data=buf, file_name="nba_2026_net_ratings.xlsx",
        mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        use_container_width=True)

    st.divider()
    st.markdown("""
### 2026 Bracket – Playoff parit

**Länsi:**
| Sija | Joukkue | vs | Sija | Joukkue |
|------|---------|----|----|---------|
| 1 | Oklahoma City Thunder | vs | 8 | Play-In voittaja |
| 4 | Los Angeles Lakers | vs | 5 | Houston Rockets |
| 3 | Denver Nuggets | vs | 6 | Minnesota Timberwolves |
| 2 | San Antonio Spurs | vs | 7 | Play-In voittaja |
| Play-In 7 | Phoenix Suns | vs | 8 | Portland Trail Blazers |
| Play-In 9 | LA Clippers | vs | 10 | Golden State Warriors |

**Itä:**
| Sija | Joukkue | vs | Sija | Joukkue |
|------|---------|----|----|---------|
| 1 | Detroit Pistons | vs | 8 | Play-In voittaja |
| 4 | Cleveland Cavaliers | vs | 5 | New York Knicks |
| 3 | Atlanta Hawks | vs | 6 | Toronto Raptors |
| 2 | Boston Celtics | vs | 7 | Play-In voittaja |
| Play-In 7 | Philadelphia 76ers | vs | 8 | Orlando Magic |
| Play-In 9 | Charlotte Hornets | vs | 10 | Miami Heat |

### Net Rating -lähteet
- **NBA.com** → Stats → Teams → Advanced → Net Rating
- **Basketball Reference** → leagues/NBA_2026_ratings.html
- **Cleaning the Glass** → Adjusted Net Rating (paras vaihtoehto)
""")
